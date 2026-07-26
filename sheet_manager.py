import os
import time
import sqlite3
import glob
import pandas as pd
from datetime import datetime

try:
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side
    )
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class SheetManager:
    COLORS = {
        "header_bg" : "1E40AF",
        "header_fg" : "FFFFFF",
        "published"  : "D1FAE5",
        "failed"     : "FEE2E2",
        "pending"    : "FEF3C7",
        "ready"      : "DBEAFE",
        "row_alt"    : "F8FAFC",
    }

    def __init__(
        self,
        db_path     = "fb_automation_v3.db",
        output_file = "FB_Empire_Live_Tracker.xlsx",
    ):
        self.db_path     = os.path.join(os.path.dirname(__file__), db_path)
        self.output_file = os.path.join(
            os.path.dirname(__file__), output_file
        )
        self._max_backups = 3

    def _fetch_data(self, ui_log=print):
        for attempt in range(5):
            try:
                conn = sqlite3.connect(
                    self.db_path, timeout=20, check_same_thread=False
                )
                conn.execute("PRAGMA journal_mode=WAL")

                df_ws = pd.read_sql_query("""
                    SELECT
                        profile_name   AS "Workspace",
                        target_fb_url  AS "Facebook Page URL",
                        chrome_path    AS "Chrome Profile Path",
                        warmup_enabled AS "Warmup ON",
                        offset_enabled AS "Random Offset ON",
                        added_on       AS "Created On"
                    FROM workspaces ORDER BY added_on DESC
                """, conn)

                df_vid = pd.read_sql_query("""
                    SELECT
                        id             AS "ID",
                        workspace_name AS "Workspace",
                        title          AS "Title",
                        status         AS "Status",
                        schedule_time  AS "Scheduled Time",
                        source_url     AS "Source URL",
                        local_path     AS "Local File Path",
                        added_on       AS "Added On"
                    FROM page_videos ORDER BY added_on DESC
                """, conn)

                conn.close()
                return df_ws, df_vid

            except sqlite3.OperationalError as e:
                if "locked" in str(e).lower() and attempt < 4:
                    time.sleep(2)
                    continue
                return pd.DataFrame(), pd.DataFrame()
            except Exception:
                return pd.DataFrame(), pd.DataFrame()

        return pd.DataFrame(), pd.DataFrame()

    def _rotate_backup(self, ui_log=print):
        if not os.path.exists(self.output_file):
            return
        try:
            import shutil
            ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
            bak  = self.output_file.replace(".xlsx", f"_bak_{ts}.xlsx")
            shutil.copy2(self.output_file, bak)

            pattern = self.output_file.replace(".xlsx", "_bak_*.xlsx")
            backups = sorted(glob.glob(pattern))
            for old in backups[:max(0, len(backups) - self._max_backups)]:
                try:
                    os.remove(old)
                except Exception:
                    pass
        except Exception as e:
            ui_log(f"[!] Backup warning: {e}")

    def _format_sheet(self, ws, df, sheet_type="default"):
        if not OPENPYXL_AVAILABLE or df.empty:
            return

        c            = self.COLORS
        header_fill  = PatternFill("solid", fgColor=c["header_bg"])
        header_font  = Font(bold=True, color=c["header_fg"], size=11)
        header_align = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        thin   = Side(border_style="thin", color="CBD5E1")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for cell in ws[1]:
            cell.fill      = header_fill
            cell.font      = header_font
            cell.alignment = header_align
            cell.border    = border
        ws.row_dimensions[1].height = 30

        for row_idx, row in enumerate(
            ws.iter_rows(min_row=2, max_row=ws.max_row), start=2
        ):
            row_color = None
            if sheet_type == "videos":
                for i, cell in enumerate(row):
                    if ws.cell(1, i + 1).value == "Status":
                        s = str(cell.value or "").upper()
                        if "PUBLISHED" in s:
                            row_color = c["published"]
                        elif "FAILED" in s:
                            row_color = c["failed"]
                        elif "READY" in s:
                            row_color = c["ready"]
                        elif "PENDING" in s:
                            row_color = c["pending"]
                        break

            for cell in row:
                cell.border    = border
                cell.alignment = Alignment(vertical="center")
                if row_color:
                    cell.fill = PatternFill("solid", fgColor=row_color)
                elif row_idx % 2 == 0:
                    cell.fill = PatternFill(
                        "solid", fgColor=c["row_alt"]
                    )

        for col_idx, col in enumerate(
            ws.iter_cols(min_row=1, max_row=ws.max_row), start=1
        ):
            max_len = max(
                (len(str(cell.value or "")) for cell in col), default=10
            )
            ws.column_dimensions[
                get_column_letter(col_idx)
            ].width = min(50, max(12, max_len + 4))

        ws.freeze_panes = "A2"

    def sync_to_excel(self, ui_log=print) -> bool:
        ui_log("[*] Syncing to Excel tracker...")
        df_ws, df_vid = self._fetch_data(ui_log)

        if df_ws.empty and df_vid.empty:
            ui_log("[!] No data to export.")
            return False

        def cnt(kw):
            if df_vid.empty:
                return 0
            return int(
                df_vid["Status"].str.upper()
                .str.contains(kw, na=False).sum()
            )

        df_summary = pd.DataFrame({
            "Metric": [
                "Total Videos", "⏳ Pending",
                "✅ Ready", "🎉 Published",
                "❌ Failed", "Last Synced",
            ],
            "Value": [
                len(df_vid), cnt("PENDING"),
                cnt("READY"), cnt("PUBLISHED"),
                cnt("FAILED"),
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            ],
        })

        # FIX: Use _tmp.xlsx not .tmp
        tmp_path = self.output_file.replace(".xlsx", "_tmp.xlsx")

        try:
            with pd.ExcelWriter(tmp_path, engine="openpyxl") as writer:
                df_vid.to_excel(
                    writer, sheet_name="📹 Video Tracker", index=False
                )
                df_ws.to_excel(
                    writer, sheet_name="📁 Workspaces", index=False
                )
                df_summary.to_excel(
                    writer, sheet_name="📊 Summary", index=False
                )

            if OPENPYXL_AVAILABLE:
                from openpyxl import load_workbook
                wb = load_workbook(tmp_path)
                self._format_sheet(
                    wb["📹 Video Tracker"], df_vid, "videos"
                )
                self._format_sheet(
                    wb["📁 Workspaces"], df_ws, "default"
                )
                self._format_sheet(
                    wb["📊 Summary"], df_summary, "default"
                )
                wb.save(tmp_path)

            self._rotate_backup(ui_log)

            if os.path.exists(self.output_file):
                os.remove(self.output_file)
            os.rename(tmp_path, self.output_file)

            ui_log(f"[+] ✅ Excel updated: {os.path.basename(self.output_file)}")
            ui_log(
                f"    📊 {len(df_vid)} videos | "
                f"{cnt('PUBLISHED')} published | "
                f"{cnt('PENDING')} pending"
            )
            return True

        except PermissionError:
            ui_log("[-] Excel file is open. Please close it!")
            if os.path.exists(tmp_path):
                try:
                    os.remove(tmp_path)
                except Exception:
                    pass
            return False

        except Exception as e:
            ui_log(f"[-] Excel sync failed: {type(e).__name__}: {e}")
            if os.path.exists(tmp_path):
                try:
                    os.remove(tmp_path)
                except Exception:
                    pass
            return False

    def export_to_csv(self, output_csv="fb_export.csv", ui_log=print):
        _, df_vid = self._fetch_data(ui_log)
        if df_vid.empty:
            ui_log("[!] No data.")
            return False
        try:
            path = os.path.join(os.path.dirname(__file__), output_csv)
            df_vid.to_csv(path, index=False, encoding="utf-8-sig")
            ui_log(f"[+] CSV: {path}")
            return True
        except Exception as e:
            ui_log(f"[-] CSV failed: {e}")
            return False

    def get_stats(self, ui_log=print) -> dict:
        _, df_vid = self._fetch_data(ui_log)
        if df_vid.empty:
            return {}

        def cnt(kw):
            return int(
                df_vid["Status"].str.upper()
                .str.contains(kw, na=False).sum()
            )

        return {
            "total"    : len(df_vid),
            "pending"  : cnt("PENDING"),
            "ready"    : cnt("READY"),
            "published": cnt("PUBLISHED"),
            "failed"   : cnt("FAILED"),
        }


# Global singleton
sheet_manager = SheetManager()